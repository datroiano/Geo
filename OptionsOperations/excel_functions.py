import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, numbers
from pathlib import Path
import os
import subprocess


def save_to_excel(df, avg_return, std_dev):
    # Get the path to the Downloads folder
    downloads_path = Path.home() / "Downloads"

    # Create the full path for the Excel file
    excel_file_path = downloads_path / "simulated_data.xlsx"

    df['entry_time'] = pd.to_datetime(df['entry_time']).dt.strftime('%H:%M:%S')
    df['exit_time'] = pd.to_datetime(df['exit_time']).dt.strftime('%H:%M:%S')

    # Set up ExcelWriter and specify a larger column width for all columns
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        workbook = writer.book
        sheet = writer.sheets['Sheet1']

        for column in df:
            column_width = max(df[column].astype(str).map(len).max(), len(column))
            col_idx = df.columns.get_loc(column)
            col_letter = get_column_letter(col_idx + 1)
            writer.sheets['Sheet1'].column_dimensions[
                col_letter].width = column_width * 1.2  # Adjusted for better readability

        # Set time format for the first two columns
        for col in ['A', 'B']:  # Assuming 'A' and 'B' are the first two columns
            col_letter = col
            sheet.column_dimensions[col_letter].number_format = 'HH:MM:SS'

        # Set percentage format for the last column
        last_col = get_column_letter(len(df.columns))
        for cell in sheet[last_col]:
            cell.number_format = numbers.FORMAT_PERCENTAGE_00

        # Adjust width for statstical data columns
        sheet.column_dimensions['J'].width = 15  # Adjust the width for column H
        sheet.column_dimensions['K'].width = 15  # Adjust the width for column I

        # Writing AVG RETURN and STD DEV as percentages with 2 decimal places
        sheet['J2'] = 'AVG RETURN'
        sheet['J3'].number_format = '0.00%'  # Format cell H3 as percentage with 2 decimal places
        sheet['J3'] = avg_return
        sheet['K2'] = 'STD DEV'
        sheet['K3'].number_format = '0.00%'  # Format cell I3 as percentage with 2 decimal places
        sheet['K3'] = std_dev

        # Center align all cells
        for row in sheet.iter_rows(min_row=1, max_row=(len(df) + 1)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f"Excel file saved to: {excel_file_path}")


def open_recent_download():
    # Get the path to the Downloads folder
    downloads_path = Path.home() / "Downloads"

    # Get a list of files in the Downloads folder along with their modification times
    files = [(f, f.stat().st_mtime) for f in downloads_path.iterdir() if f.is_file()]

    # Sort the list of files by modification time in descending order
    files.sort(key=lambda x: x[1], reverse=True)

    # Ensure there's at least one file in the list
    if len(files) > 0:
        most_recent_file = files[0][0]  # Get the most recent file (first in the sorted list)

        # Open the most recent file using the default system application
        try:
            if os.name == 'nt':  # For Windows
                os.startfile(most_recent_file)
            elif os.name == 'posix':  # For Linux, macOS, etc.
                subprocess.run(['xdg-open', most_recent_file])
        except Exception as e:
            print(f"Failed to open the file: {e}")
    else:
        print("No files found in the Downloads folder.")

